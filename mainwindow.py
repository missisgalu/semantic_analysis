# mainwindow.py

# PyQt5
from PyQt5.QtWidgets import QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QTextEdit, QMessageBox, QFileDialog, QTabWidget, QApplication, QDialog, QSizePolicy, QScrollArea
from PyQt5.QtCore import Qt, QUrl, pyqtSignal
from PyQt5.QtWidgets import QTextBrowser
from PyQt5.QtGui import QTextCursor, QDesktopServices
from PyQt5.QtWidgets import QTableWidgetItem
from PyQt5.QtCore import QThread
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QProgressBar

from scipy.spatial.distance import cosine

from semantic_analysis_tab import SemanticAnalysisTab

# matplotlib
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

# auth_utils
from auth_utils import get_auth_url, paste_auth_code, get_access_token, show_confirmation_window

# word2vec_utils
from word2vec_utils import preprocess_text, tokenize_text, lemmatize_tokens, normalize_text, train_word2vec_model, remove_stopwords

# gensim
from gensim.models import Word2Vec

# scipy
from scipy.spatial.distance import cosine

# QTableWidget
from PyQt5.QtWidgets import QTableWidget

# openpyxl
import openpyxl
from openpyxl.styles import Alignment

import docx2txt
import PyPDF4
from docx import Document
from PyPDF4 import PdfFileReader

import csv
import logging
import requests
import os

# Класс окна авторизации
class AuthWindow(QWidget):

    # Сигнал успешной авторизации
    auth_success = pyqtSignal(str)

    def __init__(self):
        super().__init__()

        # Настройка интерфейса окна
        self.setWindowTitle("Авторизация в приложении")
        self.resize(480, 200)

        self.setWindowIcon(QIcon('logo.png')) 

        self.auth_link_label = QLabel()
        self.auth_link_label.setText(f'<a href="{get_auth_url()}" style="text-decoration: none; color: black; background-color: #48D1CC; padding: 10px; border-radius: 5px;">|  Нажмите для авторизации  |</a>')
        self.auth_link_label.setOpenExternalLinks(True)
        self.auth_link_label.setAlignment(Qt.AlignCenter)
        self.auth_link_label.setFixedWidth(460)

        self.auth_code_entry = QLineEdit()
        self.auth_code_entry.setPlaceholderText("Введите код авторизации (буквы после https://example.com/vacancy-parser?code=)")
        self.auth_code_entry.setStyleSheet("QLineEdit { padding: 10px; }")

        self.auth_code_button = QPushButton("Ввести код")
        self.auth_code_button.clicked.connect(self.enter_auth_code)
        self.auth_code_button.setStyleSheet("QPushButton { background-color: #4CAF50; color: black; padding: 10px; }")

        layout = QVBoxLayout()
        layout.addWidget(self.auth_link_label)
        layout.addWidget(self.auth_code_entry)
        layout.addWidget(self.auth_code_button)

        self.setLayout(layout)

    # Проверка валидности кода авторизации
    def is_auth_code_valid(self, auth_code):
        if auth_code:
            return True
        return False

    # Обработка нажатия кнопки авторизации 
    def enter_auth_code(self):

        # Получение кода из поля ввода
        auth_code = self.auth_code_entry.text()

        try:
            # Проверка кода
            if not self.is_auth_code_valid(auth_code):
                raise AuthError("Неверный код авторизации")
      
            access_token = get_access_token(auth_code)
    
        except AuthError as e:
            QMessageBox.critical(self, "Ошибка авторизации", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", "Не удалось получить токен")  
            logging.exception(e)
            return

        # Отправка сигнала с токеном
        if access_token:
            self.auth_success.emit(access_token)
            self.close()

# Класс ошибки авторизации
class AuthError(Exception):
    pass

# Класс для вкладки поиска вакансий
class SearchVacancyTab(QWidget):

    def __init__(self):
        super().__init__()

        self.filtered_vacancies = [] 
        self.ACCESS_TOKEN = ""
        self.progress_bar = QProgressBar()

        self.setWindowTitle("Поиск вакансий")
        self.resize(600, 400)
        
        self.keywords_entry = QLineEdit()
        self.keywords_entry.setPlaceholderText("Введите слова через запятую")
        
        self.excluded_keywords_entry = QLineEdit()
        self.excluded_keywords_entry.setPlaceholderText("Введите слова-исключения")
        
        self.search_button = QPushButton("Найти вакансии")
        self.search_button.clicked.connect(self.search_vacancies)
        self.search_button.setStyleSheet("QPushButton { background-color: #42A5F5; color: black; padding: 10px; }")

        self.result_text = QTextBrowser()
        self.result_text.setReadOnly(True)
        self.result_text.setOpenExternalLinks(True)
        self.result_text.setPlaceholderText("Результаты поиска")
        self.result_text.anchorClicked.connect(self.open_link)

        self.save_results_vacancy_button = QPushButton("Сохранить в Excel")
        self.save_results_vacancy_button.clicked.connect(self.save_results_vacancy)
        self.save_results_vacancy_button.setStyleSheet("QPushButton { background-color: #4CAF50; color: black; padding: 10px; }")

        self.train_word2vec_model_button = QPushButton("Обучить модель")
        self.train_word2vec_model_button.clicked.connect(self.show_train_model)
        self.train_word2vec_model_button.setStyleSheet("QPushButton { background-color: #cc80ff; color: black; padding: 10px; }")

        main_layout = QVBoxLayout()
        main_layout.addWidget(self.keywords_entry)
        main_layout.addWidget(self.excluded_keywords_entry)
        main_layout.addWidget(self.search_button)
        main_layout.addWidget(self.result_text)
        main_layout.addWidget(self.progress_bar)
        main_layout.addWidget(self.save_results_vacancy_button)
        main_layout.addWidget(self.train_word2vec_model_button)

        self.setLayout(main_layout)
    
    def train_word2vec_model(self, texts, progress_callback=None):

        model = train_word2vec_model(texts, progress_callback)
        
        return model

    def show_train_model(self):

        text = self.result_text.toPlainText()
        
          # проверяем что найдено вакансий
        if len(self.filtered_vacancies) == 0:
            QMessageBox.warning(self, "Ошибка", "Нельзя обучить модель, т.к. найдено 0 вакансий")  
            return

        text = self.result_text.toPlainText()
        
        if not text:
            QMessageBox.warning(self, "Ошибка", "Сначала найдите вакансии")
            return

        self.progress_bar.show()

        model = self.train_word2vec_model([text], self.on_train_progress)  

        if not model:
            QMessageBox.warning(self, "Ошибка", "Не удалось обучить модель")
            return

        words = list(model.wv.index_to_key)

        dialog = TrainModelDialog(words, model)
        dialog.finished.connect(self.on_train_finished)
        dialog.exec_()
  
    def on_train_finished(self):
        self.progress_bar.setValue(0)
        self.progress_bar.hide()
    
    def on_train_progress(self, progress):
        self.progress_bar.setValue(progress)
    
    # Метод для сохранения найденных вакансий в Excel
    def save_results_vacancy(self):
        try:
            if self.filtered_vacancies:
                file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить результаты поиска", "", "Файлы Excel (*.xlsx)")

                if file_path:
                    workbook = openpyxl.Workbook()
                    worksheet = workbook.active
                    worksheet.title = "Результаты поиска"

                    # Настройка ширины столбцов
                    worksheet.column_dimensions['A'].width = 25
                    worksheet.column_dimensions['B'].width = 150
                    worksheet.column_dimensions['C'].width = 40

                    # Выравнивание по верхнему краю
                    worksheet.legacy_horizontal_alignment = 'left'
                    worksheet.legacy_vertical_alignment = 'top'

                    # Включить перенос текста для всего диапазона ячеек
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True)

                    # Заголовки столбцов
                    worksheet.append(["Название вакансии", "Описание", "Ссылка"])

                    # Добавление данных о вакансиях
                    for vacancy in self.filtered_vacancies:
                        vacancy_name = str(vacancy['name'])
                        vacancy_description = str(vacancy.get('snippet', 'Описание отсутствует'))
                        vacancy_url = str(vacancy['alternate_url'])
                        # Создание кликабельной ссылки
                        vacancy_link = f'<a href="{vacancy_url}">{vacancy_name}</a>'
                        
                        worksheet.append([vacancy_name, vacancy_description, vacancy_url])

                    workbook.save(file_path)
                    QMessageBox.information(self, "Сохранение результатов поиска", "Результаты поиска успешно сохранены.")
                else:
                    QMessageBox.critical(self, "Ошибка", "Не указан путь для сохранения.")
            else:
                QMessageBox.critical(self, "Ошибка", "Нет результатов поиска.")
        except AttributeError:
            QMessageBox.critical(self, "Ошибка", "Сначала введите ключевые слова.")

    def open_link(self, url):
        if url.scheme() == "http" or url.scheme() == "https":
            QDesktopServices.openUrl(url)

    # Метод для поиска вакансий через API hh.ru 
    def search_vacancies(self):

        # Проверка наличия токена
        if self.ACCESS_TOKEN:
            access_token = self.ACCESS_TOKEN
            keywords = self.keywords_entry.text()
            excluded_keywords = self.excluded_keywords_entry.text()

            # Формирование URL и заголовков запроса
            url = "https://api.hh.ru/vacancies"
            headers = {"Authorization": f"Bearer {access_token}"}
            params = {"text": keywords}

            # Отправка GET запроса
            response = requests.get(url, headers=headers, params=params)

            # Обработка ответа при коде 200 
            if response.status_code == 200:
                # Парсинг ответа в JSON
                vacancies = response.json()
                # Фильтрация вакансий по ключевым словам
                filtered_vacancies = vacancies.get("items", [])

                # Дополнительная фильтрация по исключающим словам
                if excluded_keywords:
                    excluded_keywords_list = [keyword.strip() for keyword in excluded_keywords.split(",")]
                    filtered_vacancies = [vacancy for vacancy in filtered_vacancies if not any(keyword.lower() in vacancy["name"].lower() for keyword in excluded_keywords_list)]

                self.filtered_vacancies = filtered_vacancies

                self.result_text.clear()

                # Вывод результатов в интерфейс
                self.result_text.insertPlainText(f"Найдено {len(filtered_vacancies)} вакансий\n\n")
                for vacancy in filtered_vacancies:
                    self.result_text.insertPlainText(f"Название вакансии: {vacancy['name']}\n")
                    self.result_text.insertPlainText(f"Описание: {vacancy.get('snippet', 'Описание отсутствует')}\n")
                    url = vacancy["alternate_url"]
                    self.result_text.append(f'<a href="{url}">{url}</a>')
                    self.result_text.insertPlainText("\n" + "-" * 20 + "\n")
            
            # Обработка ошибок запроса
            else:
                QMessageBox.critical(self, "Ошибка при поиске вакансий", "Произошла ошибка при выполнении запроса.")
        else:
            QMessageBox.critical(self, "Ошибка авторизации", "Вы не авторизованы. Введите токен доступа.")

# Класс диалогового окна для отображения обученной модели
class TrainModelDialog(QDialog):
    def __init__(self, words, model):
        super().__init__()

        self.setWindowTitle("Обученная модель")
        self.resize(600, 400)

        # Текстовое поле для отображения слов модели
        self.words_text = QTextEdit()
        # Установка режима только для чтения
        self.words_text.setReadOnly(True)
        # Отображение списка слов
        self.words_text.setPlainText("\n".join(words))

        self.save_model_button = QPushButton("Сохранить модель")
        self.save_model_button.clicked.connect(self.save_model)
        self.save_model_button.setStyleSheet("QPushButton { background-color: #4CAF50; color: black; padding: 10px; }")

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.words_text)
        self.main_layout.addWidget(self.save_model_button)
        self.setLayout(self.main_layout)

        self.vacancy_model = model
    
    # Метод для сохранения модели
    def save_model(self):
        model_path, _ = QFileDialog.getSaveFileName(self, "Сохранить модель", "", "Файлы моделей (*.model)")
        if model_path and self.vacancy_model is not None:
            self.vacancy_model.save(model_path)
            QMessageBox.information(self, "Сохранение модели", "Модель успешно сохранена.")
            self.close()
        else:
            QMessageBox.critical(self, "Ошибка", "Не выбран путь для сохранения модели или модель не обучена.")
            self.close()

# Класс для вкладки обучения модели дисциплин
class DisciplinesTab(QWidget):
    def __init__(self):
        super().__init__()

        self.ACCESS_TOKEN = ""
        self.progress_bar = QProgressBar()

        self.setWindowTitle("Дисциплины")
        self.resize(600, 400)

        self.words_text = QTextEdit()
        self.words_text.setReadOnly(True)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(self.words_text)

        # Атрибут для хранения пути к папке с дисциплинами
        self.disciplines_folder_path = ""
        # Атрибут для хранения текстов дисциплин
        self.discipline_texts = []
        # Атрибут для хранения обученной модели
        self.discipline_model = None

        self.disciplines_folder_label = QLabel()
        self.disciplines_folder_label.setWordWrap(True)
        self.disciplines_folder_label.hide()

        self.select_disciplines_folder_button = QPushButton("Выбрать папку с дисциплинами")
        self.select_disciplines_folder_button.clicked.connect(self.select_disciplines_folder)
        self.select_disciplines_folder_button.setStyleSheet("QPushButton { background-color: #42A5F5; color: black; padding: 10px; }")

        self.train_model_button = QPushButton("Обучить модель")
        self.train_model_button.clicked.connect(self.train_model)
        self.train_model_button.setStyleSheet("QPushButton { background-color: #cc80ff; color: black; padding: 10px; }")

        self.save_word2vec_model_button = QPushButton("Сохранить модель Word2Vec")
        self.save_word2vec_model_button.clicked.connect(self.save_word2vec_model)
        self.save_word2vec_model_button.setStyleSheet("QPushButton { background-color: #4CAF50; color: black; padding: 10px; }")

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.select_disciplines_folder_button)
        self.main_layout.addWidget(self.disciplines_folder_label)
        self.main_layout.addWidget(self.train_model_button)
        self.main_layout.addWidget(scroll_area)
        self.main_layout.addWidget(self.progress_bar)
        self.main_layout.addWidget(self.save_word2vec_model_button)

        self.setLayout(self.main_layout)


    def show_train_model(self):

        if not self.discipline_model:
            QMessageBox.warning(self, "Ошибка", "Сначала обучите модель")
            return
        
        words = list(self.discipline_model.wv.index_to_key)

        dialog = TrainModelDialog(words, self.discipline_model)
        dialog.exec_()

    def train_model(self):

        if not self.discipline_texts:
            QMessageBox.warning(self, "Ошибка", "Нет текстов для обучения")
            return

        self.discipline_model = train_word2vec_model(self.discipline_texts)
        
        if not self.discipline_model:
            QMessageBox.warning(self, "Ошибка", "Обучение модели не удалось")
            return
            
        QMessageBox.information(self, "Успех", "Модель обучена")

        self.show_train_model()

    def select_disciplines_folder(self):
        # Открытие диалога выбора папки
        folder_path = QFileDialog.getExistingDirectory(self, "Выберите папку с дисциплинами")

        if folder_path:
            self.disciplines_folder_path = folder_path
            self.load_disciplines(self.disciplines_folder_path)
            self.disciplines_folder_label.setText(f"Выбрана папка: {self.disciplines_folder_path}")
            self.disciplines_folder_label.adjustSize()
            self.disciplines_folder_label.show()

    def load_disciplines(self, folder_path):
        supported_extensions = ['.docx', '.txt', '.pdf', '.xlsx']
        discipline_texts = []
        problematic_files = []
        # Цикл по файлам в папке
        for file_name in os.listdir(folder_path):
            # Формирование пути к файлу
            file_path = os.path.join(folder_path, file_name)
            # Проверка поддерживаемых расширений
            if os.path.isfile(file_path) and os.path.splitext(file_name)[1] in supported_extensions:
                try:
                    if file_path.endswith('.docx'):
                        with open(file_path, 'rb') as docx_file:
                            discipline_texts.append(docx2txt.process(docx_file))
                    elif file_path.endswith('.txt'):
                        with open(file_path, 'r', encoding='utf-8') as txt_file:
                            discipline_texts.append(txt_file.read())
                    elif file_path.endswith('.pdf'):
                        with open(file_path, 'rb') as pdf_file:
                            pdf_reader = PyPDF4.PdfFileReader(pdf_file)
                            for page_num in range(pdf_reader.getNumPages()):
                                page = pdf_reader.getPage(page_num)
                                discipline_texts.append(page.extractText())
                    elif file_path.endswith('.xlsx'):
                        workbook = openpyxl.load_workbook(file_path)
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            for row in sheet.iter_rows(values_only=True):
                                for cell_value in row:
                                    if cell_value:
                                        discipline_texts.append(str(cell_value))
                    else:
                        problematic_files.append(file_name)
                except Exception as e:
                    problematic_files.append(file_name)
                    logging.warning(f"Не удалось загрузить файл {file_name}: {e}")
            else:
                problematic_files.append(file_name)
        
        if problematic_files:
            QMessageBox.warning(self, "Предупреждение", f"Не удалось загрузить следующие файлы: {', '.join(problematic_files)}")

        self.discipline_texts = discipline_texts
        if not self.discipline_texts:
            QMessageBox.warning(self, "Предупреждение", "Не найдено ни одного файла нужного формата для обучения модели.")
            
        return discipline_texts

    def train_model(self):
        if self.disciplines_folder_path:
            if self.discipline_texts:
                self.progress_bar.setMaximum(len(self.discipline_texts))
                self.progress_bar.setValue(0)
                self.progress_bar.show()
                self.discipline_model = train_word2vec_model(self.discipline_texts, lambda progress: self.update_progress(progress))
                self.progress_bar.hide()
                QMessageBox.information(self, "Обучение модели", "Обучение модели завершено.")
                self.show_words_dialog()
            else:
                QMessageBox.warning(self, "Предупреждение", "Не найдено ни одного файла нужного формата для обучения модели.")
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите папку с дисциплинами.")

    def update_progress(self, progress):
        self.progress_bar.setValue(progress)
        QApplication.processEvents()

    def show_words_dialog(self):
        if self.discipline_model:
            words = list(self.discipline_model.wv.index_to_key)
            self.words_text.setPlainText("\n".join(words))
            self.words_text.adjustSize()

    def save_word2vec_model(self):
        if self.discipline_model:
            file_path = QFileDialog.getSaveFileName(self, "Сохранить модель Word2Vec", "", "Модель Word2Vec (*.model)")[0]
            if file_path:
                self.discipline_model.save(file_path)
                QMessageBox.information(self, "Сохранение модели", "Модель успешно сохранена.")
        else:
            QMessageBox.critical(self, "Ошибка", "Обучите модель Word2Vec.")

    def select_discipline_model(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите модель дисциплины", "", "Модель Word2Vec (*.model)")
        if file_path:
            self.discipline_model = Word2Vec.load(file_path)
            QMessageBox.information(self, "Загрузка модели", "Модель успешно загружена.")

# Главное окно приложения
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Атрибут для хранения модели вакансий
        self.vacancy_model = None 

        # Создание окна авторизации
        self.auth_window = AuthWindow()
        # Подключение сигнала успешной авторизации
        self.auth_window.auth_success.connect(self.handle_auth_success)

        # Атрибут для хранения токена доступа
        self.ACCESS_TOKEN = ""

        # Настройка заголовка окна
        self.setWindowTitle("Приложение семантического анализа")
        self.resize(600, 420)
        
        self.setWindowIcon(QIcon('logo.png'))

        # Создание вкладок приложения
        self.tabs = QTabWidget()
        self.search_tab = SearchVacancyTab()
        self.disciplines_tab = DisciplinesTab()
        self.semantic_analysis_tab = SemanticAnalysisTab()
        
        # Добавление вкладок в интерфейс
        self.tabs.addTab(self.search_tab, "Поиск вакансий")
        self.tabs.addTab(self.disciplines_tab, "Дисциплины")
        self.tabs.addTab(self.semantic_analysis_tab, "Семантический анализ")

        # Установка виджета вкладок как центрального в интерфейсе
        self.setCentralWidget(self.tabs)
    
    def open_auth_url(self, event):
        QDesktopServices.openUrl(QUrl(self.auth_link_label.text()))

    # Обработка успешной авторизации
    def handle_auth_success(self, access_token):
        # Открытие главного окна
        self.open_main_window(access_token)

    # Метод для открытия главного окна
    def open_main_window(self, access_token):

        # Сохранение токена доступа
        self.ACCESS_TOKEN = access_token
        # Закрытие окна авторизации
        self.auth_window.close()
        # Передача токена доступа вкладкам
        self.search_tab.ACCESS_TOKEN = access_token
        self.disciplines_tab.ACCESS_TOKEN = access_token
        self.semantic_analysis_tab.ACCESS_TOKEN = access_token
        # Отображение главного окна
        self.show()
